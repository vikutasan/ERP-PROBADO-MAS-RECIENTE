"""
Script para sembrar clientes y rutas del módulo Grandeza
desde el Excel MAYO 2026.xlsx
Ejecutar dentro del container rderico-api-dev.
"""
import openpyxl
import asyncio
from sqlalchemy import text
from core.database import AsyncSessionLocal

wb = openpyxl.load_workbook('/app/mayo2026.xlsx', data_only=True)

day_map = {
    'NUEVO LUNES': 'LUNES',
    'NUEVO MARTES': 'MARTES',
    'NUEVO MIERCOLES': 'MIERCOLES',
    'NUEVO JUEVES': 'JUEVES',
    'NUEVO VIERNES': 'VIERNES'
}

skip_labels = ['nombre', 'negocio', 'telefono', 'direccion', 'ubicacion', 'clave', 'compra']

def clean_val(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.lower() in skip_labels or s == '' or s == 'None':
        return None
    return s

# ── Paso 1: Extraer clientes únicos ──────────────────────────────────────────
all_clients = {}  # key (UPPER) -> {name, business_name, phone, address, google_maps_url, days: {day: order}}
routes_by_day = {}  # day -> [(visit_order, client_key)]

for sheet_name, day_name in day_map.items():
    ws = wb[sheet_name]
    routes_by_day[day_name] = []
    row = 11
    visit_order = 1

    while row <= ws.max_row:
        header = ws.cell(row=row, column=1).value
        if not header or 'CLIENTE' not in str(header):
            break

        nombre = clean_val(ws.cell(row=row+4, column=1).value)
        negocio = clean_val(ws.cell(row=row+7, column=1).value)
        telefono = clean_val(ws.cell(row=row+10, column=1).value)
        direccion = clean_val(ws.cell(row=row+12, column=1).value)
        ubicacion = clean_val(ws.cell(row=row+15, column=1).value)

        if nombre:
            client_key = nombre.upper().strip()
            if client_key not in all_clients:
                all_clients[client_key] = {
                    'name': nombre,
                    'business_name': negocio,
                    'phone': telefono,
                    'address': direccion,
                    'google_maps_url': ubicacion if ubicacion and 'http' in ubicacion.lower() else None,
                    'days': {}
                }
            c = all_clients[client_key]
            if not c['phone'] and telefono:
                c['phone'] = telefono
            if not c['address'] and direccion:
                c['address'] = direccion
            if not c['business_name'] and negocio:
                c['business_name'] = negocio
            if not c['google_maps_url'] and ubicacion and 'http' in str(ubicacion).lower():
                c['google_maps_url'] = ubicacion

            c['days'][day_name] = visit_order
            routes_by_day[day_name].append((visit_order, client_key))

        row += 18
        visit_order += 1


# ── Paso 2: Insertar en la base de datos ─────────────────────────────────────
async def seed():
    async with AsyncSessionLocal() as db:
        # Verificar si ya hay clientes
        result = await db.execute(text("SELECT COUNT(*) FROM grandeza_clients"))
        count = result.scalar()
        if count > 0:
            print(f"⚠️  Ya existen {count} clientes en grandeza_clients.")
            print("   Limpiando route_slots y clients para re-sembrar...")
            await db.execute(text("DELETE FROM grandeza_route_slots"))
            await db.execute(text("DELETE FROM grandeza_clients"))
            await db.commit()
            print("   ✅ Tablas limpiadas.")

        # Insertar clientes
        client_id_map = {}  # client_key -> db_id
        for client_key, data in all_clients.items():
            result = await db.execute(
                text("""
                    INSERT INTO grandeza_clients (name, business_name, phone, address, google_maps_url, active)
                    VALUES (:name, :business_name, :phone, :address, :google_maps_url, true)
                    RETURNING id
                """),
                {
                    'name': data['name'],
                    'business_name': data['business_name'],
                    'phone': data['phone'],
                    'address': data['address'],
                    'google_maps_url': data['google_maps_url'],
                }
            )
            client_id = result.scalar()
            client_id_map[client_key] = client_id
            print(f"  ✅ Cliente #{client_id}: {data['name']}" +
                  (f" ({data['business_name']})" if data['business_name'] else "") +
                  (f" | Tel: {data['phone']}" if data['phone'] else "") +
                  f" | Días: {list(data['days'].keys())}")

        await db.commit()
        print(f"\n🎉 {len(client_id_map)} clientes insertados.\n")

        # Insertar route_slots
        slot_count = 0
        for day_name, route in routes_by_day.items():
            print(f"  📅 {day_name}:")
            for visit_order, client_key in route:
                client_id = client_id_map.get(client_key)
                if client_id:
                    await db.execute(
                        text("""
                            INSERT INTO grandeza_route_slots (client_id, day_of_week, visit_order)
                            VALUES (:client_id, :day_of_week, :visit_order)
                        """),
                        {
                            'client_id': client_id,
                            'day_of_week': day_name,
                            'visit_order': visit_order,
                        }
                    )
                    client_name = all_clients[client_key]['name']
                    print(f"     #{visit_order} → {client_name}")
                    slot_count += 1

        await db.commit()
        print(f"\n🎉 {slot_count} route_slots insertados en total.")

        # Resumen final
        print("\n" + "="*60)
        print("📊 RESUMEN FINAL")
        print("="*60)
        for day in ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES']:
            count = len(routes_by_day.get(day, []))
            print(f"  {day}: {count} clientes")
        print(f"  TOTAL CLIENTES ÚNICOS: {len(client_id_map)}")
        print(f"  TOTAL SLOTS DE RUTA:   {slot_count}")
        print("="*60)

asyncio.run(seed())
