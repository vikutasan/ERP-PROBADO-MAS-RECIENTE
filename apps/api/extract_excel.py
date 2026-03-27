import openpyxl
import json
import os
from datetime import datetime

class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(self, obj)

file_path = 'temp_masas.xlsx'
if not os.path.exists(file_path):
    print(f"Error: {file_path} not found")
    exit(1)

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    data = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_data = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                sheet_data.append(list(row))
        data[sheet_name] = sheet_data

    print(json.dumps(data, indent=2, ensure_ascii=False, cls=DateTimeEncoder))
except Exception as e:
    print(f"Error reading Excel: {str(e)}")
