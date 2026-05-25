import openpyxl
import json

wb = openpyxl.load_workbook('/app/mayo2026.xlsx', data_only=True)

day_map = {
    'NUEVO LUNES': 'LUNES',
    'NUEVO MARTES': 'MARTES',
    'NUEVO MIERCOLES': 'MIERCOLES',
    'NUEVO JUEVES': 'JUEVES',
    'NUEVO VIERNES': 'VIERNES'
}

all_clients = {}
routes_by_day = {}

for sheet_name, day_name in day_map.items():
    ws = wb[sheet_name]
    routes_by_day[day_name] = []

    row = 11
    visit_order = 1
    while row <= ws.max_row:
        header = ws.cell(row=row, column=1).value
        if not header or 'CLIENTE' not in str(header):
            break

        # Client block pattern (18 rows each):
        # row+0:  CLIENTE N header
        # row+3:  'Nombre' label
        # row+4:  actual name value
        # row+6:  'Negocio' label
        # row+9:  'Telefono' label
        # row+10: actual phone value
        # row+11: 'Direccion' label
        # row+12: actual address value
        # row+14: 'UBICACION' label
        # row+15: actual ubicacion value

        nombre = ws.cell(row=row+4, column=1).value
        negocio_raw = ws.cell(row=row+7, column=1).value
        telefono = ws.cell(row=row+10, column=1).value
        direccion = ws.cell(row=row+12, column=1).value
        ubicacion = ws.cell(row=row+15, column=1).value

        # Filter out label text
        skip_labels = ['nombre', 'negocio', 'telefono', 'direccion', 'ubicacion', 'clave', 'compra']

        def clean_val(v, skip_list=skip_labels):
            if v is None:
                return None
            s = str(v).strip()
            if s.lower() in skip_list or s == '' or s == 'None':
                return None
            return s

        nombre_clean = clean_val(nombre)
        negocio_clean = clean_val(negocio_raw)
        telefono_clean = clean_val(telefono)
        direccion_clean = clean_val(direccion)
        ubicacion_clean = clean_val(ubicacion)

        if nombre_clean:
            client_key = nombre_clean.upper()
            if client_key not in all_clients:
                all_clients[client_key] = {
                    'name': nombre_clean,
                    'business_name': negocio_clean,
                    'phone': telefono_clean,
                    'address': direccion_clean,
                    'google_maps_url': ubicacion_clean if ubicacion_clean and 'http' in ubicacion_clean.lower() else None,
                    'days': {}
                }
            # Update with any new data found
            c = all_clients[client_key]
            if not c['phone'] and telefono_clean:
                c['phone'] = telefono_clean
            if not c['address'] and direccion_clean:
                c['address'] = direccion_clean
            if not c['business_name'] and negocio_clean:
                c['business_name'] = negocio_clean
            if not c['google_maps_url'] and ubicacion_clean and 'http' in str(ubicacion_clean).lower():
                c['google_maps_url'] = ubicacion_clean

            c['days'][day_name] = visit_order
            routes_by_day[day_name].append((visit_order, nombre_clean))

        row += 18
        visit_order += 1

print("=== CLIENTES ENCONTRADOS ===")
for name, data in all_clients.items():
    print(json.dumps({
        'key': name,
        'name': data['name'],
        'business_name': data['business_name'],
        'phone': data['phone'],
        'address': data['address'],
        'google_maps_url': data['google_maps_url'],
        'days': data['days']
    }, ensure_ascii=False))

print(f"\nTOTAL: {len(all_clients)} clientes unicos")
print()
for day, route in routes_by_day.items():
    print(f"{day} ({len(route)} clientes): {[r[1] for r in route]}")
