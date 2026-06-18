import openpyxl
import json
import sys

wb = openpyxl.load_workbook(r'C:\Users\servidor1\Downloads\ENERO 2026.xlsx', data_only=True)

print("=== HOJAS DISPONIBLES ===")
print(wb.sheetnames)

clients = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== HOJA: {sheet_name} ===")
    print(f"Filas: {ws.max_row}, Columnas: {ws.max_column}")
    
    # Buscar bloques de clientes
    # Pattern: look for rows with "Clave", "Nombre", "Negocio", "Telefono", "Direccion"
    for row in range(1, ws.max_row + 1):
        cell_a = str(ws.cell(row=row, column=1).value or '').strip()
        cell_b = str(ws.cell(row=row, column=2).value or '').strip()
        
        # Detect client header rows like "CLIENTE 1", "CLIENTE 2", etc.
        if 'CLIENTE' in cell_a.upper() and any(c.isdigit() for c in cell_a):
            client_num = cell_a.strip()
            print(f"\n--- {client_num} (row {row}) ---")
            
            # Read the next several rows for client data
            data = {}
            for offset in range(1, 15):  # Look ahead up to 15 rows
                r = row + offset
                if r > ws.max_row:
                    break
                label_a = str(ws.cell(row=r, column=1).value or '').strip().lower()
                val_b = ws.cell(row=r, column=2).value
                val_c = ws.cell(row=r, column=3).value
                
                if 'clave' in label_a:
                    data['clave'] = str(val_b or '').strip()
                elif 'nombre' in label_a:
                    data['nombre'] = str(val_b or '').strip()
                elif 'negocio' in label_a:
                    data['negocio'] = str(val_b or '').strip()
                elif 'telefono' in label_a or 'teléfono' in label_a:
                    data['telefono'] = str(val_b or '').strip()
                elif 'direccion' in label_a or 'dirección' in label_a:
                    data['direccion'] = str(val_b or '').strip()
                elif 'ubicacion' in label_a or 'ubicación' in label_a:
                    # Check for hyperlink
                    cell = ws.cell(row=r, column=1)
                    if cell.hyperlink:
                        data['google_maps_url'] = cell.hyperlink.target
                    elif val_b:
                        data['google_maps_url'] = str(val_b).strip()
                    # Also check column B for hyperlink
                    cell_b_obj = ws.cell(row=r, column=2)
                    if cell_b_obj.hyperlink:
                        data['google_maps_url'] = cell_b_obj.hyperlink.target
                elif 'cliente' in label_a.lower() and any(c.isdigit() for c in label_a):
                    break  # Next client block
            
            if data.get('nombre'):
                key = data['nombre'].upper()
                if key not in clients or len(str(data)) > len(str(clients.get(key, {}))):
                    clients[key] = {
                        'sheet': sheet_name,
                        **data
                    }
                    print(f"  Nombre: {data.get('nombre')}")
                    print(f"  Negocio: {data.get('negocio')}")
                    print(f"  Telefono: {data.get('telefono')}")
                    print(f"  Direccion: {data.get('direccion')}")
                    print(f"  Maps: {data.get('google_maps_url', 'N/A')}")

# Also check for hyperlinks in column A for Google Maps
print("\n\n=== HYPERLINKS ENCONTRADOS ===")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column + 1, 5)):
            cell = ws.cell(row=row, column=col)
            if cell.hyperlink:
                print(f"  Sheet={sheet_name} Row={row} Col={col}: {cell.hyperlink.target}")

print("\n\n=== RESUMEN: TODOS LOS CLIENTES ÚNICOS ===")
for i, (key, data) in enumerate(sorted(clients.items()), 1):
    print(f"\n{i}. {data.get('nombre', 'N/A')}")
    print(f"   Negocio: {data.get('negocio', 'N/A')}")
    print(f"   Telefono: {data.get('telefono', 'N/A')}")
    print(f"   Direccion: {data.get('direccion', 'N/A')}")
    print(f"   Maps: {data.get('google_maps_url', 'N/A')}")
    print(f"   Hoja origen: {data.get('sheet', 'N/A')}")
